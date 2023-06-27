import openpyxl
import requests
import re
from os import getcwd,path
import time, json
from parsel import Selector
from datetime import datetime

wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet["A1"] = "SKU"
data_sheet["B1"] = "Brand"
data_sheet["C1"] = "Mfg"
data_sheet["D1"] = "Description"
data_sheet["E1"] = "Features"
data_sheet["F1"] = "Specifications"
data_sheet["G1"] = "Fabric"
data_sheet["H1"] = "Includes"
data_sheet["I1"] = "Ingredients"
data_sheet["J1"] = "Brand Link"

file_path = getcwd() 
# file_name = input('Enter Name-----> : ')
# output_filename = f"{file_path}\\{file_name}.xlsx"  
max_row = 2
item = {}

skus = ['40392-001-M4W6','11MWHRH','40114-0YA-W4','40112-83Z-W4','40114-76G-W4','40003-6VP-M6W8','40009-610-M6W8','40129-610-M11','40123-2BO-W6','40115-9CT-W4']
for sku in skus:
    print('------>>>>> ',sku)
    res = requests.get(f'https://www.heydude.com/search?q={sku}')
    res_xpath = Selector(text=res.text)
    value = res_xpath.xpath('//div[@data-section-load="product-grid"]//div[@data-tile-wrapper]/a/@href').get('')
    if value:
        url_formation = 'https://www.heydude.com'+ value
        sku_response = requests.get(url_formation)
        sku_xpath = Selector(text=sku_response.text)
        if sku in sku_response.text:
            id = re.findall(r'ProductID\:\s*(.*?),',sku_response.text)[0]
            feature_xpath = sku_xpath.xpath(f'//div[@data-show="{id}"]')
            title = sku_xpath.xpath('//h1/text()').get('').strip()
            brand = sku_xpath.xpath('//meta[@content="Brand"]/following-sibling::meta[1]/@content').get('').strip()
            Mfg = ''
            description = sku_xpath.xpath(f'//meta[@name="twitter:description"]/@content').get('').strip()
            empty_list = []
            if  sku_xpath.xpath(f'//div[@data-show="{id}"]//*[contains(text(),"Features & Benefits")]/parent::b/parent::p//text()'):
                features = [i.replace('\u200b','').strip() for i in sku_xpath.xpath(f'//div[@data-show="{id}"]//*[contains(text(),"Features & Benefits")]/parent::b/parent::p//text()').getall()]
                for i in features:                
                    if i!='':
                        empty_list.append(i)
            elif feature_xpath.xpath(f'//div[@data-show="{id}"]//*[contains(text(),"Specs:")]/parent::p/following-sibling::ul/li//text()|//div[@data-show="{id}"]//*[contains(text(),"features:")]/parent::p/following-sibling::ul/li//text()'):
                features = [i.replace('\u200b','').strip() for i in feature_xpath.xpath(f'//div[@data-show="{id}"]//*[contains(text(),"Specs:")]/parent::p/following-sibling::ul/li//text()|//div[@data-show="{id}"]//*[contains(text(),"features:")]/parent::p/following-sibling::ul/li//text()').getall()]
                for i in features:
                    if i!='':
                        empty_list.append(i)
            # else:
            #     features = 'New Tag'
            breakpoint()
            Features= ''
            Specifications = ''
            Fabric = ''
            Includes = ''
            Ingredients = ''
            Brand_link = url_formation
            data_sheet.cell(row =max_row, column =1).value =sku
            data_sheet.cell(row =max_row, column =2).value =brand
            data_sheet.cell(row =max_row, column =3).value =''
            data_sheet.cell(row =max_row, column =4).value =description
            data_sheet.cell(row =max_row, column =5).value =('\n').join(features)
            data_sheet.cell(row =max_row, column =6).value =''
            data_sheet.cell(row =max_row, column =7).value =''
            data_sheet.cell(row =max_row, column =8).value =''
            data_sheet.cell(row =max_row, column =9).value =''
            data_sheet.cell(row =max_row, column =10).value =url_formation
            images = ['https:'+i.strip() for i in sku_xpath.xpath('//div[@data-thumbnails-wrapper]//following-sibling::div//div[@data-product-images]//img/@src').getall()]
            for num, image_set in enumerate(images,1):
                if f'Images {num}' not in item.keys():
                    col_value = data_sheet.max_column+1
                    item[f'Images {num}'] = col_value
                    data_sheet.cell(row =1, column =col_value).value = f'Images {num}'
                    data_sheet.cell(row =max_row, column =col_value).value = image_set
                else:
                    data_sheet.cell(row =max_row, column =item[f'Images {num}']).value = image_set
            max_row+=1
            try:
                wb.save(f"{file_path}\\Heydude_Input_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
            except:
                print(('Please Close the file'))
                wb.save(f"{file_path}\\Heydude_Input_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
            
