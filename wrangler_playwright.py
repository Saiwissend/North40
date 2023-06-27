# import os
# import re
# import csv
# import json
# import pandas as pd
# from datetime import datetime
import openpyxl
import requests
import json
import re
from os import getcwd,path
import time, json
from parsel import Selector
from datetime import datetime
import json
from parsel import Selector
from playwright.sync_api import sync_playwright


wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet["A1"] = "SKU"
data_sheet["B1"] = "Brand"
data_sheet["C1"] = "Mfg"
data_sheet["D1"] = "Description"
data_sheet["E1"] = "Features"
data_sheet["F1"] = "Specifications"
data_sheet["G1"] = "Fabric"
data_sheet["H1"] = "Includes"
data_sheet["I1"] = "Ingredients"
data_sheet["J1"] = "Brand Link"


file_path = getcwd() 
# file_name = input('Enter Name-----> : ')
# output_filename = f"{file_path}\\{file_name}.xlsx"  
max_row = 2
item = {}
skus = ["33MWXAB","112315083","112333881","MT1STST","13MWBSW","14MWZGH","112334569","13MGSSW","3W194GY","112327612"]
with sync_playwright() as playwright:
    webkit = playwright.firefox
    browser = webkit.launch(headless=False)
    context = browser.new_context(java_script_enabled=True)
    page = context.new_page()
    for sku in skus:
        page.goto(f"https://www.wrangler.com/search?q={sku}",wait_until="load")
        response = Selector(text=page.content())
        value = 'https://www.wrangler.com'+response.xpath('//div[@class="product"]//a/@href').get('')
        if value:
            url_formation = value
            page.goto(url_formation,wait_until="load")
            response_sku = Selector(text=page.content())
            if sku in page.content():
                id = response_sku.xpath('//script[@type="application/ld+json"]/text()').get('')
                extracting_json = json.loads(id)
                title = response_sku.xpath('//h1/text()').get('').strip()
                brand = re.findall(r'Brand\"\,\"name\"\:\"(.*?)\"\}',page.content())[0]
                Mfg = ''
                description = ' '.join([i.strip() for i in response_sku.xpath('//div[@class="description-item-content-wrapper"]//p//text()').getall()])
                empty_list = []
                for i in response_sku.xpath('//div[@class="description-item-content-wrapper"]/following-sibling::ul/li'):
                    empty_replace = ' '.join([j.strip() for j in i.xpath('.//text()').getall()]).strip()
                    empty_list.append(empty_replace)    
                if response_sku.xpath('//*[@id="pdp-fabric"]/text()').get('').strip():
                    fabrics =response_sku.xpath('//*[@id="pdp-fabric"]/text()').get('').strip()
                else:
                    fabrics = ''
                Features= ''
                Specifications = ''
                Fabric = ''
                Includes = ''
                Ingredients = ''
                Brand_link = url_formation
                data_sheet.cell(row =max_row, column =1).value =sku
                data_sheet.cell(row =max_row, column =2).value =brand
                data_sheet.cell(row =max_row, column =3).value =''
                data_sheet.cell(row =max_row, column =4).value =description
                data_sheet.cell(row =max_row, column =5).value =('\n').join(empty_list)
                data_sheet.cell(row =max_row, column =6).value =''
                data_sheet.cell(row =max_row, column =7).value =fabrics
                data_sheet.cell(row =max_row, column =8).value =''
                data_sheet.cell(row =max_row, column =9).value =Ingredients
                data_sheet.cell(row =max_row, column =10).value =url_formation
                if response_sku.xpath('//div[@class="row no-gutters pdp-image-carousel  sceneSevenData"]/div//img/@src'):
                    images = [i.strip() for i in response_sku.xpath('//div[@class="row no-gutters pdp-image-carousel  sceneSevenData"]/div//img/@src').getall()]
                else:
                    images = [response_sku.xpath('//img[@data-image-index="0.0"]/@src').get('')]
                for num, image_set in enumerate(images,1):
                    if f'Images {num}' not in item.keys():
                        col_value = data_sheet.max_column+1
                        item[f'Images {num}'] = col_value
                        data_sheet.cell(row =1, column =col_value).value = f'Images {num}'
                        data_sheet.cell(row =max_row, column =col_value).value = image_set
                    else:
                        data_sheet.cell(row =max_row, column =item[f'Images {num}']).value = image_set
                max_row+=1
                try:
                    wb.save(f"{file_path}\\Wrangler_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
                except:
                    print(('Please Close the file'))
                    wb.save(f"{file_path}\\Wrangler_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
                

    