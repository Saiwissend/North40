import openpyxl
import requests
import json
import re
import urllib.parse
from os import getcwd,path
import time, json
from parsel import Selector
from datetime import datetime

wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet["A1"] = "Brand"
data_sheet["B1"] = "Mfg"
data_sheet["C1"] = "Description"
data_sheet["D1"] = "Features"
data_sheet["E1"] = "Specifications"
data_sheet["F1"] = "Fabric"
data_sheet["G1"] = "Includes"
data_sheet["H1"] = "Ingredients"
data_sheet["I1"] = "Brand Link"

file_path = getcwd() 

max_row = 2
item = {}

sku = '112325402'
# for i in range(1,10):
#     encode_url=urllib.parse.quote((f'https://www.mscdirect.com/product/details/74077660'))
#     link = f'http://api.scrape.do/?token=c2e5fc7c447f445da3ec04d6c9061cbb0bd58642e75&url={encode_url}'
#     res = requests.get(link)
#     print(res)
#     breakpoint()

res = requests.get(f'https://www.wrangler.com/search?q={sku}')
res_xpath = Selector(text=res.text)
value = res_xpath.xpath('//div[@class="product"]//a/@href').get('')
if value:
    breakpoint()
    url_formation = value
    sku_response = requests.get(url_formation)
    sku_xpath = Selector(text=sku_response.text)
    # breakpoint()
    if sku in sku_response.text:
        id = sku_xpath.xpath('//script[@type="application/ld+json"]/text()').get('')
        extracting_json = json.loads(id)
        title = sku_xpath.xpath('//h1/text()').get('').strip()
        brand = re.findall(r'Brand\"\,\"name\"\:\"(.*?)\"\}',sku_response.text)[0]
        Mfg = ''
        description = ' '.join([i.strip() for i in sku_xpath.xpath('//div[@class="r-product-desc"]//p//text()').getall()])
        empty_list = []
        for i in sku_xpath.xpath('//div[@class="r-details-features"]//ul/li'):
            empty_replace = ' '.join([j.strip() for j in i.xpath('.//text()').getall()]).strip()
            empty_list.append(empty_replace)    
        if re.findall(r'\"hrB2bCharacteristics\"\:\"Fabric:s*(.*?)\,\"hrB2bDescription\"',sku_response.text):
            fabrics = re.findall(r'\"hrB2bCharacteristics\"\:\"Fabric:s*(.*?)\,\"hrB2bDescription\"',sku_response.text)[0].strip()
        else:
            fabrics = ''
        Features= ''
        Specifications = ''
        Fabric = ''
        Includes = ''
        Ingredients = sku_xpath.xpath('//span[contains(text(),"Composition")]/p/text()').get('').strip()
        Brand_link = url_formation
        data_sheet.cell(row =max_row, column =1).value =brand
        data_sheet.cell(row =max_row, column =2).value =''
        data_sheet.cell(row =max_row, column =3).value =description
        data_sheet.cell(row =max_row, column =4).value =('\n').join(empty_list)
        data_sheet.cell(row =max_row, column =5).value =''
        data_sheet.cell(row =max_row, column =6).value =fabrics
        data_sheet.cell(row =max_row, column =7).value =''
        data_sheet.cell(row =max_row, column =8).value =Ingredients
        data_sheet.cell(row =max_row, column =9).value =url_formation
        images = [i.strip() for i in sku_xpath.xpath('//div[@class="r-productimages for-desktop "]//div[@class="product-thumbnails-nav-carousel"]/ul//li//picture//img/@src').getall()]
        for num, image_set in enumerate(images,1):
            updated_image = image_set.replace('medium-large','hi-res')
            if f'Images {num}' not in item.keys():
                col_value = data_sheet.max_column+1
                item[f'Images {num}'] = col_value
                data_sheet.cell(row =1, column =col_value).value = f'Images {num}'
                data_sheet.cell(row =max_row, column =col_value).value = updated_image
            else:
                data_sheet.cell(row =max_row, column =item[f'Images {num}']).value = updated_image
        max_row+=1
        try:
            wb.save(f"{file_path}\\roxy_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
        except:
            print(('Please Close the file'))
            wb.save(f"{file_path}\\roxy_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
        
