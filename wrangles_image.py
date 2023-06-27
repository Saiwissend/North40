import openpyxl
import requests
import json
import re
import os
from os import getcwd,path
import time, json
from parsel import Selector
from datetime import datetime
import json
from parsel import Selector
from io import BytesIO
from urllib.request import urlopen
from zipfile import ZipFile
from playwright.sync_api import sync_playwright


wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet.title='Image'
data_sheet["A1"] = "Brand"
data_sheet["B1"] = "SKU"

ws2 = wb.create_sheet('Product_collection')
ws2['A1'] = 'title'   
ws2['B1'] = 'brand'   
ws2['C1'] = 'item_number' 
ws2['D1'] = 'SKU'   
ws2['E1'] = 'response.url'

max_row =2
spec_sheet = 2
item = {}

file_path = getcwd() 
inupt_files = f"{file_path}\\Zip Files"
if not os.path.isdir(inupt_files):
	os.mkdir(inupt_files) 
skus = ['09MWGMS','MS1041D','112339757','MS71519','112330683']
for sku in skus:
  print('SKU---------->  ',sku)
  url = f"https://site.wranglertraderspass.com/tp_search.php/search?search_term={sku}"

  payload = 'fs_csrf_secure_token=b243cb7442ef06a4c43b6ab751512d4a6108fdc6c6c44b8f9ae13bc751feb5e9&user_email=sheenac%40csww.net&user_password=Media%23123&site_id=1'
  headers = {
  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
  'Accept-Language': 'en-US,en;q=0.9',
  'Cache-Control': 'max-age=0',
  'Connection': 'keep-alive',
  'Content-Type': 'application/x-www-form-urlencoded',
  'Cookie': 'PHPSESSID=9mlne16b0o62tqof5l98utmj6h; _ga=GA1.1.816149303.1686200620; _ga_49LWGDBM7R=GS1.1.1686200619.1.0.1686200632.0.0.0; user_encrypted_data=%7B%22ct%22%3A%22leirzGyJGg3d-WxP4MeuSoGpYl10sZqTq5zMqqn87CGRAAJknW7AGgIymQA5eehf-OUVvUyeF3rsNewh2bKhQHjxiUu0B3uja2ceeelOczGAV6zmzUX4RI4oyPxhELku4SMW07T6bMROqusIeifhdqDcR7JfMviOkao17VEYjKWsZqITN6TURnPtPGpueRcU6d5xNWd8Ggfln1YYFchWLKS2WblB-hNfMPtEpcEZ5PZdqeaaOSE_GeEihNtuF8jEI2enWWLi88vnt-a71lVT0k44AtbDXkHQstBFDMCOOo86c3TN9CLgwFn2cd5YN5PiZn_-XV1NZYtKejbZg2udDwrCkoUlx0eZ576lNLpTl0XyA3_nqhyBdOnBpvzfavuNAddDVilAPokW-cNToJJryPp1N_4mEcAbubNHdC0mgXN-HaF3JKiGc-znIc5jQgQ08bDNwpT4jNHAQj3NuihwvIXOcWT6j3qOCj2nqpKXyhGxxKENilZGih8GxVTORAE250rP1miOxNE2JRne2hNGi1lsqassQPbx0j2B_fDzF3-5wq5Gi5KYuFSvSAGB0p5IA_wZPZ8Od7ga-1bY-WQGp6tHJek7c7SaKaqCP4AXliqytni47gQr7MJju7-LISmoanTVvYdAdygoSKcUAcL_D6vBgMR6iOVmNZlTBpR-Uv334XoPjV6cBb0pQz-qWaHxMxuk1dM0oUj2F1ju6apQqa5HBJzaYS9alEAjAk27GPel04n5NyJkZTtYbk2qZVuLqf854ljX-WjvXfehQFZyLjD2F-P8UMbgenGMHuwLRxeiGv-L0n25hx_5xC9ycjDLFdejJUN79zvObSQCHZg8o40vEmCG9FPWssIF-5IAGmDDqSn2jZw-dVtrYbT6PBBUcX5x00Bb3WLcvbWFeooblhNGy0JxdYlE6sLwCbRR-btY0sX8arPtu6EM9lrNMmj5UipXy6lv-EzSiVOTnZVwAVgQpfcESB8n9zNYGBZ2Pj7WhAqYgUHpCsZ_w6CQaAUnzsuaovWaBojw7wiuwDo55K6jUE-toVy_BOEuGIfiEpYURZKpLtBaG-YeeI45cscwvF2q_XvG-k00wVU5fKD-eVSkiwmsWYBZBU_yuF1agcWUVv8VG-CpuM9RJFFudaHJEy2JMOxoVOBYjvFc1bQq17rHjdeSNdWUCnfvaHU81oB1HqwI8sfl2y4MWXd5D9Z7Q-1Bk21c1v5bEwkplXk0z0OofmqfeobYsFk6-CpR7WxH8qm3jBizetkWonR435cAl9F26OZYhaeVX2N3oraCzx3U_xd_DkOYUHMXE6Y-v0GP5Inj2vMr3lANAL8t1rzIcKwcN9NO3i9nVYVhYHM-FA%22%2C%22iv%22%3A%2218b592627da1db88fea81c8ab877481d%22%2C%22s%22%3A%227476bded33fb938c%22%7D',
  'Origin': 'https://site.wranglertraderspass.com',
  'Referer': 'https://site.wranglertraderspass.com/',
  'Sec-Fetch-Dest': 'document',
  'Sec-Fetch-Mode': 'navigate',
  'Sec-Fetch-Site': 'same-origin',
  'Sec-Fetch-User': '?1',
  'Upgrade-Insecure-Requests': '1',
  'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
  'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
  'sec-ch-ua-mobile': '?0',
  'sec-ch-ua-platform': '"Windows"'
  }
  response = requests.request("POST", url, headers=headers, data=payload)
  response_xpath = Selector(text=response.text)
  urls = ['https://site.wranglertraderspass.com'+i.replace("download_zip('",'').replace(",'download');",'').strip() for i in response_xpath.xpath('//div[@class="pkb-section"]//button[contains(@onclick,"download_zip")]/@onclick').getall()]
  if urls:
    for num, image_set in enumerate(urls,1):
      if f'Images {num}' not in item.keys():
        col_value = data_sheet.max_column+1
        item[f'Images {num}'] = col_value
        data_sheet.cell(row =max_row, column =1).value ='Wrangler'
        data_sheet.cell(row =max_row, column =2).value =sku
        data_sheet.cell(row =1, column =col_value).value = f'Images {num}'
        data_sheet.cell(row =max_row, column =col_value).value = image_set
      else:
        data_sheet.cell(row =max_row, column =item[f'Images {num}']).value = image_set
    max_row+=1
    for product in  response_xpath.xpath('//div[@class="pkb-item itemSKU "]/a'):
        product_url = 'https://site.wranglertraderspass.com'+ product.xpath('./@href').get('')
        if product_url:
          product_ping = requests.get(product_url,headers=headers)
          product_selector = Selector(text=product_ping.text)
          title = product_selector.xpath('//*[contains(text(),"Item Name: ")]/text()').get('').replace('Item Name: ','').strip()
          brand = product_selector.xpath('//*[contains(text(),"Brand: ")]/text()').get('').replace('Brand: ','').strip()
          item_number = product_selector.xpath('//*[contains(text(),"Item Number: ")]/text()').get('').replace('Item Number: ','').strip()
          ws2.cell(row =spec_sheet, column =1).value =title
          ws2.cell(row =spec_sheet, column =2).value =brand
          ws2.cell(row =spec_sheet, column =3).value =item_number
          ws2.cell(row =spec_sheet, column =4).value =sku
          ws2.cell(row =spec_sheet, column =5).value =response.url
          spec_sheet+=1
  else:
      print('Not Available in site-----> ',sku)
      data_sheet.cell(row =max_row, column =1).value ='Wrangler'
      data_sheet.cell(row =max_row, column =2).value =sku
      max_row+=1
  try:
      wb.save(f"{file_path}\\Wrangler_Image_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
  except:
      print(('Please Close the file'))
      wb.save(f"{file_path}\\Wrangler_Image_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
     