import openpyxl
import requests
import json
import re
from os import getcwd,path
import time, json
from parsel import Selector
from datetime import datetime

wb = openpyxl.Workbook()
data_sheet = wb.active
data_sheet["A1"] = "SKU"
data_sheet["B1"] = "Brand"
data_sheet["C1"] = "Mfg"
data_sheet["D1"] = "Description"
data_sheet["E1"] = "Features"
data_sheet["F1"] = "Specifications"
data_sheet["G1"] = "Fabric"
data_sheet["H1"] = "Includes"
data_sheet["I1"] = "Ingredients"
data_sheet["J1"] = "Brand Link"

file_path = getcwd() 
# file_name = input('Enter Name-----> : ')
# output_filename = f"{file_path}\\{file_name}.xlsx"  
max_row = 2
item = {}
skus = ['197095099351','197095097357','197095099412','195718915989','195718918492','195718917563','195718917839','195718931002','195718915774','195718917457']
for sku in skus:
    res = requests.get(f'https://www.roxy.com/search?q={sku}')
    res_xpath = Selector(text=res.text)
    value = res_xpath.xpath('//a[@aria-labelledby="View Product"]/@href').get('')
    if value:
        url_formation = value
        sku_response = requests.get(url_formation)
        sku_xpath = Selector(text=sku_response.text)
        # breakpoint()
        if sku in sku_response.text:
            id = sku_xpath.xpath('//script[@type="application/ld+json"]/text()').get('')
            extracting_json = json.loads(id)
            title = sku_xpath.xpath('//h1/text()').get('').strip()
            brand = re.findall(r'Brand\"\,\"name\"\:\"(.*?)\"\}',sku_response.text)[0]
            Mfg = ''
            description = ' '.join([i.strip() for i in sku_xpath.xpath('//div[@class="r-product-desc"]//p//text()').getall()])
            empty_list = []
            for i in sku_xpath.xpath('//div[@class="r-details-features"]//ul/li'):
                empty_replace = ' '.join([j.strip() for j in i.xpath('.//text()').getall()]).strip()
                empty_list.append(empty_replace)    
            if re.findall(r'\"hrB2bCharacteristics\"\:\"Fabric:s*(.*?)\,\"hrB2bDescription\"',sku_response.text):
                fabrics = re.findall(r'\"hrB2bCharacteristics\"\:\"Fabric:s*(.*?)\,\"hrB2bDescription\"',sku_response.text)[0].strip()
            else:
                fabrics = ''
            # featu = sku_xpath.xpath('//div[@class="r-details-features"]//ul/li//text()').getall()        
            Features= ''
            Specifications = ''
            Fabric = ''
            Includes = ''
            Ingredients = sku_xpath.xpath('//span[contains(text(),"Composition")]/p/text()').get('').strip()
            Brand_link = url_formation
            data_sheet.cell(row =max_row, column =1).value =sku
            data_sheet.cell(row =max_row, column =2).value =brand
            data_sheet.cell(row =max_row, column =3).value =''
            data_sheet.cell(row =max_row, column =4).value =description
            data_sheet.cell(row =max_row, column =5).value =('\n').join(empty_list)
            data_sheet.cell(row =max_row, column =6).value =''
            data_sheet.cell(row =max_row, column =7).value =fabrics
            data_sheet.cell(row =max_row, column =8).value =''
            data_sheet.cell(row =max_row, column =9).value =Ingredients
            data_sheet.cell(row =max_row, column =10).value =url_formation
            # images = extracting_json['image']
            images = [i.strip() for i in sku_xpath.xpath('//div[@class="r-productimages for-desktop "]//div[@class="product-thumbnails-nav-carousel"]/ul//li//picture//img/@src|//div[@class="r-productimages for-desktop vertical-video--enabled"]//div[@class="product-thumbnails-nav-carousel"]/ul//li//picture//img/@src').getall()]
            # breakpoint()
            for num, image_set in enumerate(images,1):
                updated_image = image_set.replace('medium-large','hi-res')
                if f'Images {num}' not in item.keys():
                    col_value = data_sheet.max_column+1
                    item[f'Images {num}'] = col_value
                    data_sheet.cell(row =1, column =col_value).value = f'Images {num}'
                    data_sheet.cell(row =max_row, column =col_value).value = updated_image
                else:
                    data_sheet.cell(row =max_row, column =item[f'Images {num}']).value = updated_image
            max_row+=1
            try:
                wb.save(f"{file_path}\\Roxy_Input_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
            except:
                print(('Please Close the file'))
                wb.save(f"{file_path}\\Roxy_Input_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
            
